#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_data.py — Gera o snapshot de dados da dashboard a partir da planilha.

Como usar:
  1. No Google Sheets: Arquivo > Fazer download > Microsoft Excel (.xlsx)
  2. Salve o arquivo como  Financeiro.xlsx  nesta mesma pasta.
  3. Rode:  python build_data.py
  4. Ele gera  data.json  (snapshot) e  data.js  (mesmo conteúdo, para abrir
     o index.html com duplo clique, sem servidor).

Nada sai do seu computador. É só leitura local do arquivo.
"""

import json
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta a biblioteca openpyxl. Rode:  pip install openpyxl")

ARQ = Path(__file__).with_name("Financeiro.xlsx")

# Colunas que são totais/derivados, não contas de verdade.
COLS_RESUMO = {
    "investimentos", "dinheiro", "patrimonio", "patrimônio", "delta mensal",
    "delta %", "entrada", "saída", "saida", "investimento", "% investimento",
    "% meta investimento", "total", "mês", "mes", "", "local",
}


def norm(s):
    return (str(s).strip().lower()) if s is not None else ""


def num(v):
    """Converte célula em float; devolve None se não for número."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(".", "").replace(",", ".")
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None


def ym(d):
    """Normaliza uma data para 'AAAA-MM'."""
    if isinstance(d, (datetime, date)):
        return f"{d.year:04d}-{d.month:02d}"
    return None


def classe_do_tipo(tipo, conta):
    """Agrupa cada conta em uma classe de ativo legível."""
    t = norm(tipo)
    c = norm(conta)
    if "bitcoin" in c or "cripto" in t:
        return "Cripto"
    if t in ("", "dinheiro"):
        return "Caixa & reserva"
    mapa = {
        "renda fixa": "Renda fixa", "cdb": "CDB / renda fixa",
        "fi": "Fundos de invest. (FI)",
        "fii's": "FIIs", "fiis": "FIIs", "ações": "Ações", "acoes": "Ações",
        "etf": "ETFs", "bdr": "BDRs", "td": "Tesouro Direto",
        "previdência": "Previdência", "previdencia": "Previdência",
    }
    return mapa.get(t, tipo.strip() if tipo else "Outros")


def main():
    if not ARQ.exists():
        sys.exit(f"Não encontrei {ARQ.name}. Exporte a planilha em .xlsx com esse nome.")

    wb = openpyxl.load_workbook(ARQ, data_only=True)
    out = {"meta": {}, "kpis": {}, "metas": [], "fluxo": {}, "investimentos": {}}

    out["meta"] = {
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "fonte": ARQ.name,
    }

    # ---- DASH: KPIs e metas -------------------------------------------------
    if "Dash" in wb.sheetnames:
        ws = wb["Dash"]
        rot = {}
        for r in range(1, min(60, ws.max_row) + 1):
            k = norm(ws.cell(r, 1).value)
            if k:
                rot.setdefault(k, ws.cell(r, 2).value)
        out["kpis"] = {
            "patrimonio": num(rot.get("patrimonio")) or num(rot.get("patrimonio ")),
            "investido": num(rot.get("investimento")),
            "idade": num(rot.get("idade")),
            "gasto_anual": abs(num(rot.get("gasto anual")) or 0) or None,
        }
        hoje = rot.get("hoje:")
        out["meta"]["hoje"] = (hoje.strftime("%d/%m/%Y")
                               if isinstance(hoje, (datetime, date)) else None)
        for r in range(1, min(60, ws.max_row) + 1):
            lab = ws.cell(r, 1).value
            val = num(ws.cell(r, 2).value)
            if lab and "investido" in norm(lab) and val is not None:
                out["metas"].append({"nome": str(lab).strip(), "progresso": val})

    # ---- Config_cat: mapa categoria -> natureza (fonte de verdade) ----------
    cat_natureza = {}
    if "Config_cat" in wb.sheetnames:
        wc = wb["Config_cat"]
        for r in range(2, wc.max_row + 1):
            c = norm(wc.cell(r, 1).value)
            n = norm(wc.cell(r, 2).value)
            if c:
                cat_natureza[c] = n

    # ---- FLUXO DE CAIXA: entradas/saídas reais ------------------------------
    ws = wb["Fluxo de caixa"]
    hdr = {norm(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    cData = hdr.get("data", 1)
    cValor = hdr.get("valor (r$)", 3)
    cCat = hdr.get("categoria", 4)
    cNat = hdr.get("natureza", 7)

    meses = {}          # 'AAAA-MM' -> {entradas, saidas}
    cat_saida = {}      # (mes, categoria) -> valor
    cat_saida_12m = {}
    fontes_12m = {}
    hoje = datetime.now()

    for r in range(2, ws.max_row + 1):
        cat = (ws.cell(r, cCat).value or "Sem categoria")
        cat = str(cat).strip()
        # Natureza efetiva: o mapa Config_cat manda; a coluna é fallback.
        nat = cat_natureza.get(norm(cat)) or norm(ws.cell(r, cNat).value)
        if nat != "real":
            continue
        v = num(ws.cell(r, cValor).value)
        d = ws.cell(r, cData).value
        m = ym(d)
        if v is None or m is None:
            continue
        b = meses.setdefault(m, {"entradas": 0.0, "saidas": 0.0})
        if v >= 0:
            b["entradas"] += v
            fontes_12m[cat] = fontes_12m.get(cat, 0.0) + v if _within12(d, hoje) else fontes_12m.get(cat, 0.0)
        else:
            b["saidas"] += -v
            cat_saida[(m, cat)] = cat_saida.get((m, cat), 0.0) + (-v)
            if _within12(d, hoje):
                cat_saida_12m[cat] = cat_saida_12m.get(cat, 0.0) + (-v)

    meses_ord = sorted(meses)
    out["fluxo"]["meses"] = [
        {"mes": m, "entradas": round(meses[m]["entradas"], 2),
         "saidas": round(meses[m]["saidas"], 2),
         "resultado": round(meses[m]["entradas"] - meses[m]["saidas"], 2)}
        for m in meses_ord
    ]
    if meses_ord:
        ult = meses_ord[-1]
        cats = sorted(((c, v) for (m, c), v in cat_saida.items() if m == ult),
                      key=lambda x: -x[1])
        out["fluxo"]["ultimo_mes"] = {
            "mes": ult,
            "entradas": round(meses[ult]["entradas"], 2),
            "saidas": round(meses[ult]["saidas"], 2),
            "resultado": round(meses[ult]["entradas"] - meses[ult]["saidas"], 2),
            "categorias_saida": [{"categoria": c, "valor": round(v, 2)} for c, v in cats],
        }
    out["fluxo"]["categorias_saida_12m"] = [
        {"categoria": c, "valor": round(v, 2)}
        for c, v in sorted(cat_saida_12m.items(), key=lambda x: -x[1])
    ]
    out["fluxo"]["entradas_fontes_12m"] = [
        {"fonte": c, "valor": round(v, 2)}
        for c, v in sorted(fontes_12m.items(), key=lambda x: -x[1]) if v > 0
    ]

    # ---- POSIÇÃO CONSOLIDADA: da aba fiel Patrimonio_real -------------------
    # Cols individuais de conta (2..N) têm a classe na linha 4 ("tipo").
    # Há também colunas de subtotal ("Patrimonio", "Investimentos") e um bloco
    # de subtotais por classe à direita — que reconciliam com o Dash.
    ws = wb["Patrimonio_real"]
    linha_atual = None
    for r in range(ws.max_row, 1, -1):
        if ym(ws.cell(r, 1).value):
            linha_atual = r
            break

    # localizar colunas especiais pelo nome do cabeçalho (linha 1)
    col_por_nome = {}
    for c in range(2, ws.max_column + 1):
        col_por_nome.setdefault(norm(ws.cell(1, c).value), c)
    c_patrim = col_por_nome.get("patrimonio") or col_por_nome.get("patrimônio")

    # Colunas de conta = têm uma classe em row4 e NÃO são de resumo.
    classes_label = {"renda fixa", "fii's", "fiis", "ações", "acoes", "td",
                     "cripto", "dinheiro", "bdr", "previdência", "previdencia",
                     "fi", "cdb", "etf"}
    por_classe, por_conta = {}, {}
    subtotais_classe = {}
    for c in range(2, ws.max_column + 1):
        nome = ws.cell(1, c).value
        tipo = ws.cell(4, c).value
        val = num(ws.cell(linha_atual, c).value)
        n_nome, n_tipo = norm(nome), norm(tipo)
        # Bloco de subtotais por classe (à direita da coluna "Patrimonio").
        if c > (c_patrim or 0) and n_nome in classes_label and n_tipo == n_nome:
            if val:
                subtotais_classe[classe_do_tipo(tipo, "")] = \
                    subtotais_classe.get(classe_do_tipo(tipo, ""), 0.0) + val
            continue
        if n_nome in COLS_RESUMO:
            continue
        # É uma conta individual.
        if val:
            classe = classe_do_tipo(tipo, nome)
            por_classe[classe] = por_classe.get(classe, 0.0) + val
            por_conta[str(nome).strip()] = por_conta.get(str(nome).strip(), 0.0) + val

    # Preferimos os subtotais oficiais (reconciliam com o Dash); se não houver,
    # caímos para a soma por conta.
    classe_final = subtotais_classe if subtotais_classe else por_classe
    total_pat = sum(classe_final.values()) or 1
    out["investimentos"]["por_classe"] = [
        {"classe": k, "valor": round(v, 2), "pct": round(100 * v / total_pat, 1)}
        for k, v in sorted(classe_final.items(), key=lambda x: -x[1])
    ]
    out["investimentos"]["por_conta"] = [
        {"conta": k, "valor": round(v, 2)}
        for k, v in sorted(por_conta.items(), key=lambda x: -x[1])
    ]
    out["investimentos"]["total_patrimonio"] = round(sum(classe_final.values()), 2)

    # ---- Evolução do patrimônio: coluna "Patrimonio" mês a mês --------------
    serie = []
    if c_patrim:
        for r in range(2, ws.max_row + 1):
            m = ym(ws.cell(r, 1).value)
            val = num(ws.cell(r, c_patrim).value)
            if m and val:
                serie.append({"mes": m, "valor": round(val, 2)})
    out["investimentos"]["evolucao_patrimonio"] = serie

    # ---- grava ---------------------------------------------------------------
    Path("data.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    Path("data.js").write_text("window.FIN_DATA = " + json.dumps(out, ensure_ascii=False) + ";",
                               encoding="utf-8")
    print("OK: data.json e data.js gerados.")
    print(f"  Patrimônio (Dash):        R$ {out['kpis'].get('patrimonio'):,.2f}")
    print(f"  Soma das classes:         R$ {out['investimentos']['total_patrimonio']:,.2f}")
    print(f"  Meses de fluxo de caixa:  {len(out['fluxo']['meses'])}")
    print(f"  Classes de ativo:         {len(out['investimentos']['por_classe'])}")
    print(f"  Pontos de evolução:       {len(out['investimentos']['evolucao_patrimonio'])}")


def _within12(d, hoje):
    if not isinstance(d, (datetime, date)):
        return False
    dd = datetime(d.year, d.month, d.day) if isinstance(d, date) and not isinstance(d, datetime) else d
    delta = (hoje.year - dd.year) * 12 + (hoje.month - dd.month)
    return 0 <= delta < 12


def _find_row(ws, label, col=1, scan_cols=1):
    label = norm(label)
    for r in range(1, min(ws.max_row, 60) + 1):
        for c in range(1, scan_cols + 1) if scan_cols > 1 else [col]:
            if norm(ws.cell(r, c).value) == label:
                return r
    return None


if __name__ == "__main__":
    main()
